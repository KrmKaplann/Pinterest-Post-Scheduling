import json
import os
import undetected_chromedriver as uc
import time
from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import locale
import pyautogui

from datetime import datetime


def Driver():
    options = uc.ChromeOptions()
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--start-maximized")
    driver = uc.Chrome(options=options)
    return driver

def Chrome_Tab(driver, link):
    driver.get(link)
    driver.maximize_window()


path = "<path_to_SocialMediaAccounts.xlsx>"  # Path to the Excel file containing social media accounts
workbook = openpyxl.load_workbook(path)
DefaultPageSheet = workbook["Page1"]

SocialMediaName = "Pinterest"

StartColumn = 5
while True:
    Finder = DefaultPageSheet.cell(2, StartColumn).value
    if Finder == SocialMediaName:
        break
    StartColumn += 1

TotalAccountListGmail = []
DefaultRow = 4

while True:
    Start = DefaultPageSheet.cell(DefaultRow, StartColumn).value
    if Start is None:
        break
    TotalAccountListGmail.append([Start, DefaultRow - 3])
    DefaultRow += 1

# Filter out items marked with '-'
TotalAccountListGmail = [item for item in TotalAccountListGmail if item[0] != '-']

TotalAccountList = []
DefaultRow = 4

for account_info in TotalAccountListGmail:
    index = account_info[1]  # Index value added to the first list
    Start = DefaultPageSheet.cell(index + 3, 2).value  # Calculate DefaultRow value from the index value
    TotalAccountList.append(Start)

print(TotalAccountList)


# Get the starting index from the user
starting_index = int(input("Enter the last completed index for Pinterest accounts (between 0 and {}): ".format(len(TotalAccountList)-1)))

# Loop starting from the starting index
TotalAccountList = TotalAccountList[starting_index:]
TotalAccountListGmail = TotalAccountListGmail[starting_index:]

path = "<path_to_PinterestAccounts.xlsx>"  # Path to the Excel file containing Pinterest accounts
workbook = openpyxl.load_workbook(path)
AllPagesWorksheet = workbook.sheetnames

for OneItem in TotalAccountListGmail:

    driver = Driver()
    link = "https://www.pinterest.com"

    Chrome_Tab(driver, link)

    JsonFileName = DefaultPageSheet.cell(2, StartColumn).value + OneItem[0]+".json"
    with open(f"<path_to_API_folder>/{JsonFileName}", "r") as file:  # Path to the folder containing API JSON files
        cookies = json.load(file)

    for cookie in cookies:
        driver.add_cookie(cookie)

    time.sleep(3)

    driver.refresh()

    StartRow = 5
    PostDict = {}

    AllAccountsList = AllPagesWorksheet
    print(AllAccountsList)

    time.sleep(2)
    driver.get("https://www.pinterest.com/pin-builder/")

    tabs = driver.window_handles
    first_tab = tabs[0]
    driver.switch_to.window(first_tab)

    time.sleep(2)
    if StartRow == 5:
        for i in range(3):
            time.sleep(1)
            pyautogui.hotkey('ctrl', '-')
    time.sleep(1)

    starting_index = int(input("Enter the last completed index for Pinterest accounts (between 0 and {}): ".format(len(AllAccountsList) - 1)))

    # Loop starting from the starting index
    AllAccountsList = AllAccountsList[starting_index:]

    for OneAccountName in AllAccountsList:
        StartRow = 5
        AllPostList = []
        while True:
            SpecialWorkSheet = workbook[OneAccountName]
            PostName = SpecialWorkSheet["B" + str(StartRow)].value
            StartRow += 1
            if PostName is None:
                break
            AllPostList.append(PostName)

        time.sleep(2)
        driver.get("https://www.pinterest.com/pin-builder/")
        time.sleep(2)

        print(AllPostList)
        AllPostDefault = AllPostList
        starting_index = int(input("Enter the last completed index for content (between 0 and {}): ".format(len(AllPostList) - 1)))

        # Loop starting from the starting index
        AllPostList = AllPostList[starting_index:]

        for StartRow, PostOne in enumerate(AllPostDefault, start=5):

            SpecialWorkSheet = workbook[OneAccountName]
            PostName = SpecialWorkSheet["B" + str(StartRow)].value
            Description = SpecialWorkSheet["C" + str(StartRow)].value
            WebAddress = SpecialWorkSheet["H" + str(StartRow)].value
            Category = SpecialWorkSheet["I" + str(StartRow)].value

            Date = SpecialWorkSheet["D" + str(StartRow)].value
            Date = datetime.strptime(Date, '%d.%m.%Y')
            FormattedDate = Date.strftime("%d %B %A %Y")
            Day = Date.strftime("%d")
            Month = Date.strftime("%B")
            TimeVal = str(SpecialWorkSheet["E" + str(StartRow)].value)
            TimeSplitted = TimeVal.split(":")
            Hour = str(TimeSplitted[0])
            Minutes = str(TimeSplitted[1])
            ImmediateShare = SpecialWorkSheet["F" + str(StartRow)].value

            # Create a dictionary for each post
            post_details = {
                "Description": Description,
                "Date": FormattedDate,
                "Day": Day,
                "Month": Month,
                "Time": TimeVal,
                "Hour": Hour,
                "Minutes": Minutes,
                "ImmediateShare": ImmediateShare,
                "WebAddress": WebAddress,
                "Category": Category
            }

            # If OneAccountName already exists in PostDict, retrieve the existing dictionary and add a new post
            if OneAccountName in PostDict:
                PostDict[OneAccountName][PostName] = post_details
            else:
                # If OneAccountName does not exist, create a new dictionary and add a new post
                PostDict[OneAccountName] = {PostName: post_details}

        for StartRow, PostOne in enumerate(AllPostList, start=5):
            wait = WebDriverWait(driver, 30)
            driver.implicitly_wait(10)  # Set 10 seconds implicit wait
            FileAdress = "<path_to_videos_folder>/" + OneAccountName + "/" + PostOne + ".mp4"  # Path to the folder containing video files
            time.sleep(1)
            driver.find_element(By.XPATH, "<xpath_to_upload_file_input>").send_keys(FileAdress)  # XPath to the file upload input
            time.sleep(2)

            Description = str(PostDict[OneAccountName][PostOne]["Description"])
            if "/Title " in Description:
                Title = (Description.split("/Title "))[0]
                LongDescription = (Description.split("/Title "))[1]
                driver.find_element(By.XPATH, "<xpath_to_title_input>").send_keys(Title)  # XPath to the title input field
                time.sleep(1)
                driver.find_element(By.XPATH, "<xpath_to_description_input>").send_keys(LongDescription)  # XPath to the description input field
                time.sleep(1)
            else:
                driver.find_element(By.XPATH, "<xpath_to_title_input>").send_keys(str(PostDict[OneAccountName][PostOne]["Description"]))  # XPath to the title input field
                time.sleep(1)

            if PostDict[OneAccountName][PostOne]["ImmediateShare"] == "No":
                time.sleep(2)
                driver.find_element(By.XPATH, "<xpath_to_publish_later_button>").click()  # XPath to the "Publish at a later date" button
                time.sleep(1)
            elif PostDict[OneAccountName][PostOne]["ImmediateShare"] == "Yes":
                pass

            time.sleep(3)
            CombinedDate = datetime.strptime(str(PostDict[OneAccountName][PostOne]["Date"]), '%d %B %A %Y').strftime('%d.%m.%Y')

            date_input = driver.find_element(By.XPATH, "<xpath_to_date_input>")  # XPath to the date input field
            date_input.send_keys(Keys.CONTROL + "a")
            date_input.send_keys(CombinedDate)

            time.sleep(2)
            driver.find_element(By.XPATH, "<xpath_to_time_input>").click()  # XPath to the time input field
            time.sleep(2)

            HourAndMinute = str(PostDict[OneAccountName][PostOne]["Hour"]) + ":" + str(PostDict[OneAccountName][PostOne]["Minutes"])
            # Convert the input time to a datetime object
            HourAndMinute = datetime.strptime(HourAndMinute, "%H:%M")
            # Format the time object to 12-hour format with AM/PM
            HourAndMinute = HourAndMinute.strftime("%I:%M %p")

            HourAndMinuteXPath = driver.find_element(By.XPATH, "<xpath_to_time_selection>".format(HourAndMinute))  # XPath to the specific time selection
            action = ActionChains(driver)
            action.move_to_element(HourAndMinuteXPath).perform()
            HourAndMinuteXPath.click()
            time.sleep(2)

            driver.find_element(By.XPATH, "<xpath_to_web_address_input>").send_keys(WebAddress)  # XPath to the web address input field
            time.sleep(2)

            DefaultCategoryName = driver.find_element(By.XPATH, "<xpath_to_category_button>").text  # XPath to the category button

            if PostDict[OneAccountName][PostOne]["Category"] != DefaultCategoryName:
                driver.find_element(By.XPATH, "<xpath_to_category_button>").click()  # XPath to the category button
                time.sleep(2)
                driver.find_element(By.XPATH, "<xpath_to_search_category_input>").send_keys(PostDict[OneAccountName][PostOne]["Category"])  # XPath to the search category input field
                try:
                    element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "<xpath_to_specific_category>".format(PostDict[OneAccountName][PostOne]["Category"]))))  # XPath to the specific category
                    time.sleep(1)
                    element.click()
                    time.sleep(1)
                except:
                    driver.find_element(By.XPATH, "<xpath_to_create_category_button>").click()  # XPath to the create category button
                    time.sleep(1)
                    category_name_input = driver.find_element(By.XPATH, "<xpath_to_category_name_input>")  # XPath to the category name input field
                    category_name_input.send_keys(Keys.CONTROL + "a")
                    category_name_input.send_keys(PostDict[OneAccountName][PostOne]["Category"])
                    time.sleep(1)
                    driver.find_element(By.XPATH, "<xpath_to_category_submit_button>").click()  # XPath to the category submit button
                    time.sleep(1)
            else:
                pass

            FileAdressForCover = "<path_to_cover_images_folder>/" + OneAccountName + "/" + PostOne + "-Cover" + ".jpg"  # Path to the folder containing cover images

            if os.path.exists(FileAdressForCover):
                driver.find_element(By.XPATH, "<xpath_to_cover_image_upload_input>").send_keys(FileAdressForCover)  # XPath to the cover image upload input
                time.sleep(1)

            if len(AllPostList) > StartRow - 5 + 1:
                driver.find_element(By.XPATH, "<xpath_to_batch_create_button>").click()  # XPath to the batch create button
                time.sleep(1)

            print(PostOne)


        elements = driver.find_elements(By.XPATH, "<xpath_to_save_buttons>")  # XPath to the save buttons

        for element in elements:
            element.click()
            time.sleep(20)
